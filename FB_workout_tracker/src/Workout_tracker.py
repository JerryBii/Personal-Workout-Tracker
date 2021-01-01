import openpyxl
import os
import sys
from time import sleep
import xlsxwriter
import workout

#allows user to add new exercise to worksheet
def add_new_exercise(wb, choice):
    if choice == 1:
        name = input('Please enter exercise name again: ').upper()
        new_ws = wb.create_sheet(name)
        create_worksheet(new_ws, name)
    else:
        name = input('Please enter new exercise name: ').upper()
        new_ws = wb.create_sheet(name)
        create_worksheet(new_ws, name)

    wb.save(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx')


    print('Would you like to log to this exercise?')
    loop = True
    while loop == True:
        print('1. Yes')
        print('2. No (Returns to main menu)')
        option = int(input('Enter Selection: '))
        if option == 1:
            log_existing(wb, name, option)
            loop = False
            break
        elif option == 2:
            print('Returning to main menu...')
            sleep(1)
            loop = False
            break
        else:
            print('Invalid Selection please Try Again...')
            sleep(1)

#creates a worksheet with the basic elements of an executed exercise
def create_worksheet(ws,name = ''):
    ws.cell(1, 1).value = 'Weight(lbs): '
    ws.cell(1, 2).value = 'Sets: '
    ws.cell(1, 3).value = 'Reps: '
    ws.cell(1, 4).value = 'Date: '
    if name == '':
        pass
    else:
        ws.title = name

#finds an empty row to start writing data to
def find_next_empty_row(worksheet):
    row_num = 1
    while worksheet.cell(row_num,1).value is not None:
        row_num += 1
    return row_num

#Allows user to enter in an exercise that they recently performed (other options exist if the user has entered an  unrecognized exercise)
def log_existing(workbook, name = ' ', choice = 0):
    if choice == 1:
        exercise = name
    else:
        exercise = input('Please enter the name of the exercise: ').upper()

    if exercise in workbook.sheetnames:
        current_ws = workbook[exercise]
        row_num = find_next_empty_row(current_ws)
        column_num = 1
        weight = input('Please enter the weight used (in lbs): ')
        reps = input('Please enter the reps per set: ')
        sets = input('please enter the sets performed: ')
        date = input('please enter workout date (format: MM/DD/YY): ')
        current_ws.cell(row_num,column_num).value = weight
        current_ws.cell(row_num, column_num*2).value = reps
        current_ws.cell(row_num,column_num*3).value = sets
        current_ws.cell(row_num, column_num*4).value = date
        print('Nice Job! Those are impressive numbers!')
        sleep(1)
        print('Returning to main menu...')
        sleep(1)

    else:
        question_loop = True
        print('This exercise is not in the database, would you like to add it?')
        while question_loop == True:
            print('1. Yes')
            print('2. No (Returns to main menu)')
            choice = int(input('Enter your selection: '))
            if choice == 1:
                add_new_exercise(workbook, choice)
                question_loop = False
                print('Returning to main menu...')
                sleep(1)
                break
            elif choice == 2:
                question_loop = False
                print('Returning to main menu...')
                sleep(1)
                break
            else:
                print('Invalid selection please try again...')
                sleep(1)



    workbook.save(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx')


if os.path.isfile(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx'):

    wb = openpyxl.load_workbook(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx')
    ws = wb['BENCH PRESS']
else:

    wb = openpyxl.Workbook()
    ws1 = wb['Sheet']
    ws2 = wb.create_sheet('SQUAT', 1)
    ws3 = wb.create_sheet('DEADLIFT', 2)
    create_worksheet(ws1,'BENCH PRESS')
    create_worksheet(ws2)
    create_worksheet(ws3)
    wb.save(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx')

quit = False

print('Welcome to the workout progress logger!')
while quit != True:
    print('Options: \n 1.log to existing exercise \n 2.Create new exercise \n 3.Open Excel file for manual editing \n 0.Quit')
    option = int(input('Enter your selection: '))

    if option == 1:
        log_existing(wb)
    elif option == 2:
        add_new_exercise(wb,option)
    elif option == 3:
        os.startfile(r'C:\Users\Jerry Bi\Documents\Workout Tracker\Progress.xlsx')
    elif option == 0:
        sys.exit()
    else:
        print('Invalid Selection, returning to main menu... \n')
        sleep(1)

